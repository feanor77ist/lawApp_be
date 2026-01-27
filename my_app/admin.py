from django.contrib import admin, messages
from .models import Scenario, UserEntry, UserChatHistory, Customer, Program, TrainingGroup, GroupUser, ProgramScenario, EvaluationReport
from .forms import CustomerAdminForm, TrainingGroupAdminForm, ProgramScenarioInlineFormSet, CustomUserCreationForm, CustomUserChangeForm, ScenarioAdminForm
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse, path
from django.utils.html import format_html
from django.contrib.auth.admin import UserAdmin as DefaultUserAdmin
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from admin_auto_filters.filters import AutocompleteFilter
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db.models import Avg, Sum, Q
from django.utils.safestring import mark_safe
from django.templatetags.static import static
from django.contrib.auth.forms import PasswordResetForm
from django.conf import settings
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from my_app.views import get_online_users


# Register your models here.

@admin.register(Scenario)
class ScenarioAdmin(admin.ModelAdmin):
    form = ScenarioAdminForm
    list_display = ("name", "ai_level", "user_level", "hardness", "created_at")
    list_display_links = ("name",)
    list_filter = ("ai_level", "user_level", "hardness")
    search_fields = ("name", "description")
    ordering = ("created_at",)

class UserChatHistoryInline(admin.TabularInline):
    model = UserChatHistory
    extra = 1
    fields = ['user_query', 'gpt_response', 'timestamp']
    readonly_fields = ['timestamp']

class UserEntryTrainingGroupFilter(AutocompleteFilter):
    title = 'Eğitim Grubu'
    field_name = 'training_group'

class UserEntryScenarioFilter(AutocompleteFilter):
    title = 'Eğitim Senaryosu'
    field_name = 'scenario'

class UserEntryUserFilter(AutocompleteFilter):
    title = 'Kullanıcı'
    field_name = 'user'

@admin.register(UserEntry)
class UserEntryAdmin(admin.ModelAdmin):
    list_display = ('user', 'training_group', 'scenario', 'created_at')
    search_fields = ('user__username', 'scenario__name', 'entry_id', 'training_group__name')
    list_filter = ("created_at", UserEntryTrainingGroupFilter, UserEntryScenarioFilter, UserEntryUserFilter)
    ordering = ['-created_at']
    inlines = [UserChatHistoryInline]

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    form = CustomerAdminForm
    list_display = ("name", "email", "created_at")
    filter_horizontal = ("users",)  # ManyToManyField için çoklu seçim arayüzü
    search_fields = ("name", "email")
    list_filter = ("created_at",)
    change_form_template = 'admin/my_app/customer/change_form.html'
    change_list_template = 'admin/my_app/customer/change_list.html'

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "users":
            kwargs["queryset"] = User.objects.all().order_by('-id')
        return super().formfield_for_manytomany(db_field, request, **kwargs)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_report'] = True
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        custom = [
            path(
                '<int:object_id>/customer-report/',
                self.admin_site.admin_view(self.customer_report_view),
                name=f'{app_label}_{model_name}_customer_report',
            ),
            path(
                '<int:object_id>/customer-filters/',
                self.admin_site.admin_view(self.customer_filters_view),
                name=f'{app_label}_{model_name}_customer_filters',
            ),
            path(
                '<int:object_id>/autocomplete-training-groups/',
                self.admin_site.admin_view(self.autocomplete_training_groups),
                name=f'{app_label}_{model_name}_autocomplete_training_groups',
            ),
            path(
                '<int:object_id>/autocomplete-programs/',
                self.admin_site.admin_view(self.autocomplete_programs),
                name=f'{app_label}_{model_name}_autocomplete_programs',
            ),
            path(
                '<int:object_id>/autocomplete-program-scenarios/',
                self.admin_site.admin_view(self.autocomplete_program_scenarios),
                name=f'{app_label}_{model_name}_autocomplete_program_scenarios',
            ),
            path(
                'changelist-report/',
                self.admin_site.admin_view(self.changelist_report_view),
                name=f'{app_label}_{model_name}_changelist_report',
            ),
            path(
                'changelist-filters/',
                self.admin_site.admin_view(self.changelist_filters_view),
                name=f'{app_label}_{model_name}_changelist_filters',
            ),
            path(
                'changelist-export/',
                self.admin_site.admin_view(self.changelist_export_view),
                name=f'{app_label}_{model_name}_changelist_export',
            ),
        ]
        return custom + urls

    def customer_report_view(self, request, object_id):
        """AJAX endpoint for customer report based on filters"""

        customer = self.get_object(request, object_id)
        if customer is None:
            return JsonResponse({'error': 'Customer not found'}, status=404)

        # Get multiple values (support both single and multiple selections)
        training_group_ids = request.GET.getlist('training_group')
        program_ids = request.GET.getlist('program')
        program_scenario_ids = request.GET.getlist('program_scenario')

        # Base queryset
        training_groups = TrainingGroup.objects.filter(customer=customer)
        
        if training_group_ids:
            # Filter by selected training groups
            training_groups = training_groups.filter(id__in=training_group_ids)
        if program_ids:
            # Filter by selected programs
            training_groups = training_groups.filter(program_id__in=program_ids)

        # Get all related program scenarios
        program_scenarios = ProgramScenario.objects.filter(
            program__traininggroup__customer=customer
        ).select_related('scenario', 'program').distinct()

        if program_scenario_ids:
            # Filter by selected program scenarios
            program_scenarios = program_scenarios.filter(id__in=program_scenario_ids)

        # Build report data
        report_data = []
        for ps in program_scenarios.order_by('release_date'):
            scenario = ps.scenario
            
            # Filter training groups for this scenario
            relevant_groups = training_groups.filter(program=ps.program)
            if not relevant_groups.exists():
                continue

            # Sadece katılımcıların (users) raporlarını dahil et, trainer'ları hariç tut
            participant_user_ids = []
            for group in relevant_groups:
                participant_user_ids.extend(group.users.values_list('id', flat=True))
            participant_user_ids = list(set(participant_user_ids))  # unique

            # Aggregate data across all relevant training groups (sadece katılımcılar)
            pre_reports = EvaluationReport.objects.filter(
                training_group__in=relevant_groups,
                scenario=scenario,
                type="pre",
                user_id__in=participant_user_ids,
            )
            post_reports = EvaluationReport.objects.filter(
                training_group__in=relevant_groups,
                scenario=scenario,
                type="post",
                user_id__in=participant_user_ids,
            )

            total_users = GroupUser.objects.filter(
                training_group__in=relevant_groups
            ).values('user').distinct().count()

            pre_completed = pre_reports.values('user').distinct().count()
            pre_avg = int(round(pre_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
            pre_attempt_sum = pre_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
            pre_attempt_avg = round(pre_attempt_sum / total_users, 2) if total_users > 0 else 0

            post_completed = post_reports.values('user').distinct().count()
            post_avg = int(round(post_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
            post_attempt_sum = post_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
            post_attempt_avg = round(post_attempt_sum / total_users, 2) if total_users > 0 else 0

            # Calculate improvement
            pre_scores_by_user = dict(pre_reports.values_list("user_id", "score"))
            post_scores_by_user = dict(post_reports.values_list("user_id", "score"))
            improvement_values = []
            for user_id, pre_score in pre_scores_by_user.items():
                post_score = post_scores_by_user.get(user_id)
                if post_score is None:
                    continue
                if pre_score and pre_score > 0:
                    improvement = ((post_score - pre_score) / pre_score) * 100
                else:
                    improvement = 0
                improvement_values.append(improvement)

            improvement_avg_percentage = int(round(sum(improvement_values) / len(improvement_values))) if improvement_values else None

            # Trainer isimlerini topla (tüm relevant_groups'un trainer'ları)
            trainer_names = []
            for group in relevant_groups:
                trainers = group.trainers.all()
                trainer_names.extend([trainer.get_full_name() or trainer.username for trainer in trainers])
            trainer_names = list(set(trainer_names))  # unique
            trainers_display = ", ".join(trainer_names) if trainer_names else "—"

            report_data.append({
                "scenario_name": scenario.name,
                "program_id": ps.program.id,
                "program_name": ps.program.name,
                "trainers": trainers_display,
                "release_date": ps.release_date.strftime('%d.%m.%Y') if ps.release_date else "Yayın tarihi yok",
                "pre_completed": pre_completed,
                "pre_avg": pre_avg,
                "pre_attempt_avg": pre_attempt_avg,
                "post_completed": post_completed,
                "post_avg": post_avg,
                "post_attempt_avg": post_attempt_avg,
                "improvement_avg_percentage": improvement_avg_percentage,
                "total_users": total_users,
            })

        return JsonResponse({
            'report_data': report_data,
            'customer_name': customer.name
        })

    def customer_filters_view(self, request, object_id):
        """AJAX endpoint for loading filter options"""
        from django.http import JsonResponse
        
        customer = self.get_object(request, object_id)
        if customer is None:
            return JsonResponse({'error': 'Customer not found'}, status=404)

        # Get training groups
        training_groups = TrainingGroup.objects.filter(customer=customer).select_related('program').order_by('name')
        training_group_options = [{'id': tg.id, 'name': tg.name} for tg in training_groups]

        # Get programs
        programs = Program.objects.filter(traininggroup__customer=customer).distinct().order_by('name')
        program_options = [{'id': p.id, 'name': p.name} for p in programs]

        # Get program scenarios (support multiple program IDs)
        program_ids = request.GET.getlist('program_id')
        program_scenarios = ProgramScenario.objects.filter(
            program__traininggroup__customer=customer
        ).select_related('scenario', 'program').distinct()
        
        if program_ids:
            # Filter by selected program IDs
            program_scenarios = program_scenarios.filter(program_id__in=program_ids)
        
        program_scenario_options = [
            {
                'id': ps.id,
                'name': f"{ps.program.name} | {ps.scenario.name}",
                'program_id': ps.program.id
            }
            for ps in program_scenarios.order_by('program__name', 'scenario__name')
        ]

        return JsonResponse({
            'training_groups': training_group_options,
            'programs': program_options,
            'program_scenarios': program_scenario_options,
        })

    def changelist_filters_view(self, request):
        """AJAX endpoint for loading filter options in changelist"""
        from django.http import JsonResponse
        
        # Get all customers
        customers = Customer.objects.all().order_by('name')
        customer_options = [{'id': c.id, 'name': c.name} for c in customers]

        # Get customer IDs from request
        customer_ids = request.GET.getlist('customer_id')
        
        # Get training groups (filtered by selected customers if any)
        training_groups = TrainingGroup.objects.all()
        if customer_ids:
            training_groups = training_groups.filter(customer_id__in=customer_ids)
        training_groups = training_groups.select_related('program', 'customer').order_by('name')
        training_group_options = [{'id': tg.id, 'name': tg.name, 'customer_id': tg.customer.id} for tg in training_groups]

        # Get programs (filtered by selected customers if any)
        programs = Program.objects.all()
        if customer_ids:
            programs = programs.filter(traininggroup__customer_id__in=customer_ids).distinct()
        programs = programs.order_by('name')
        program_options = [{'id': p.id, 'name': p.name} for p in programs]

        # Get program scenarios
        program_ids = request.GET.getlist('program_id')
        program_scenarios = ProgramScenario.objects.all()
        if customer_ids:
            program_scenarios = program_scenarios.filter(
                program__traininggroup__customer_id__in=customer_ids
            )
        if program_ids:
            program_scenarios = program_scenarios.filter(program_id__in=program_ids)
        program_scenarios = program_scenarios.select_related('scenario', 'program').distinct()
        
        program_scenario_options = [
            {
                'id': ps.id,
                'name': f"{ps.program.name} | {ps.scenario.name}",
                'program_id': ps.program.id
            }
            for ps in program_scenarios.order_by('program__name', 'scenario__name')
        ]

        return JsonResponse({
            'customers': customer_options,
            'training_groups': training_group_options,
            'programs': program_options,
            'program_scenarios': program_scenario_options,
        })

    def changelist_report_view(self, request):
        """AJAX endpoint for changelist report based on filters"""
        from django.http import JsonResponse

        # Get multiple values
        customer_ids = request.GET.getlist('customer')
        training_group_ids = request.GET.getlist('training_group')
        program_ids = request.GET.getlist('program')
        program_scenario_ids = request.GET.getlist('program_scenario')

        # Base queryset
        training_groups = TrainingGroup.objects.all()
        
        if customer_ids:
            training_groups = training_groups.filter(customer_id__in=customer_ids)
        if training_group_ids:
            training_groups = training_groups.filter(id__in=training_group_ids)
        if program_ids:
            training_groups = training_groups.filter(program_id__in=program_ids)

        # Get all related program scenarios
        program_scenarios = ProgramScenario.objects.all().select_related('scenario', 'program').distinct()

        if customer_ids:
            program_scenarios = program_scenarios.filter(
                program__traininggroup__customer_id__in=customer_ids
            )
        if program_scenario_ids:
            program_scenarios = program_scenarios.filter(id__in=program_scenario_ids)

        # Build report data
        report_data = []
        for ps in program_scenarios.order_by('release_date'):
            scenario = ps.scenario
            
            # Filter training groups for this scenario
            relevant_groups = training_groups.filter(program=ps.program)
            if not relevant_groups.exists():
                continue

            # Sadece katılımcıların (users) raporlarını dahil et, trainer'ları hariç tut
            participant_user_ids = []
            for group in relevant_groups:
                participant_user_ids.extend(group.users.values_list('id', flat=True))
            participant_user_ids = list(set(participant_user_ids))  # unique

            # Aggregate data across all relevant training groups (sadece katılımcılar)
            pre_reports = EvaluationReport.objects.filter(
                training_group__in=relevant_groups,
                scenario=scenario,
                type="pre",
                user_id__in=participant_user_ids,
            )
            post_reports = EvaluationReport.objects.filter(
                training_group__in=relevant_groups,
                scenario=scenario,
                type="post",
                user_id__in=participant_user_ids,
            )

            total_users = GroupUser.objects.filter(
                training_group__in=relevant_groups
            ).values('user').distinct().count()

            pre_completed = pre_reports.values('user').distinct().count()
            pre_avg = int(round(pre_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
            pre_attempt_sum = pre_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
            pre_attempt_avg = round(pre_attempt_sum / total_users, 2) if total_users > 0 else 0

            post_completed = post_reports.values('user').distinct().count()
            post_avg = int(round(post_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
            post_attempt_sum = post_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
            post_attempt_avg = round(post_attempt_sum / total_users, 2) if total_users > 0 else 0

            # Calculate improvement
            pre_scores_by_user = dict(pre_reports.values_list("user_id", "score"))
            post_scores_by_user = dict(post_reports.values_list("user_id", "score"))
            improvement_values = []
            for user_id, pre_score in pre_scores_by_user.items():
                post_score = post_scores_by_user.get(user_id)
                if post_score is None:
                    continue
                if pre_score and pre_score > 0:
                    improvement = ((post_score - pre_score) / pre_score) * 100
                else:
                    improvement = 0
                improvement_values.append(improvement)

            improvement_avg_percentage = int(round(sum(improvement_values) / len(improvement_values))) if improvement_values else None

            # Trainer isimlerini topla (tüm relevant_groups'un trainer'ları)
            trainer_names = []
            for group in relevant_groups:
                trainers = group.trainers.all()
                trainer_names.extend([trainer.get_full_name() or trainer.username for trainer in trainers])
            trainer_names = list(set(trainer_names))  # unique
            trainers_display = ", ".join(trainer_names) if trainer_names else "—"

            report_data.append({
                "scenario_name": scenario.name,
                "program_id": ps.program.id,
                "program_name": ps.program.name,
                "trainers": trainers_display,
                "release_date": ps.release_date.strftime('%d.%m.%Y') if ps.release_date else "Yayın tarihi yok",
                "pre_completed": pre_completed,
                "pre_avg": pre_avg,
                "pre_attempt_avg": pre_attempt_avg,
                "post_completed": post_completed,
                "post_avg": post_avg,
                "post_attempt_avg": post_attempt_avg,
                "improvement_avg_percentage": improvement_avg_percentage,
                "total_users": total_users,
            })

        return JsonResponse({
            'report_data': report_data,
        })

    def changelist_export_view(self, request):
        """Export changelist report (with filters) to XLSX"""
        # Collect filters
        customer_ids = request.GET.getlist('customer')
        training_group_ids = request.GET.getlist('training_group')
        program_ids = request.GET.getlist('program')
        program_scenario_ids = request.GET.getlist('program_scenario')

        # Base queryset
        training_groups = TrainingGroup.objects.all()
        if customer_ids:
            training_groups = training_groups.filter(customer_id__in=customer_ids)
        if training_group_ids:
            training_groups = training_groups.filter(id__in=training_group_ids)
        if program_ids:
            training_groups = training_groups.filter(program_id__in=program_ids)

        # Scenarios to include
        program_scenarios = ProgramScenario.objects.all().select_related('scenario', 'program').distinct()
        if customer_ids:
            program_scenarios = program_scenarios.filter(
                program__traininggroup__customer_id__in=customer_ids
            )
        if program_scenario_ids:
            program_scenarios = program_scenarios.filter(id__in=program_scenario_ids)

        # Build report rows (satır başına TrainingGroup)
        rows = []
        for ps in program_scenarios.order_by('release_date'):
            scenario = ps.scenario
            relevant_groups = training_groups.filter(program=ps.program).select_related("customer")
            if not relevant_groups.exists():
                continue

            for group in relevant_groups.order_by("name"):
                # Sadece katılımcıların (users) raporlarını dahil et, trainer'ları hariç tut
                participant_users = group.users.all()
                
                pre_reports = EvaluationReport.objects.filter(
                    training_group=group,
                    scenario=scenario,
                    type="pre",
                    user__in=participant_users,
                )
                post_reports = EvaluationReport.objects.filter(
                    training_group=group,
                    scenario=scenario,
                    type="post",
                    user__in=participant_users,
                )

                total_users = GroupUser.objects.filter(
                    training_group=group
                ).values('user').distinct().count()

                pre_completed = pre_reports.values('user').distinct().count()
                pre_avg = int(round(pre_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
                pre_attempt_sum = pre_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
                pre_attempt_avg = round(pre_attempt_sum / total_users, 2) if total_users > 0 else 0

                post_completed = post_reports.values('user').distinct().count()
                post_avg = int(round(post_reports.aggregate(avg=Avg("score"))["avg"] or 0, 2))
                post_attempt_sum = post_reports.aggregate(total=Sum("attempt_count"))["total"] or 0
                post_attempt_avg = round(post_attempt_sum / total_users, 2) if total_users > 0 else 0

                pre_scores_by_user = dict(pre_reports.values_list("user_id", "score"))
                post_scores_by_user = dict(post_reports.values_list("user_id", "score"))
                improvement_values = []
                for user_id, pre_score in pre_scores_by_user.items():
                    post_score = post_scores_by_user.get(user_id)
                    if post_score is None:
                        continue
                    if pre_score and pre_score > 0:
                        improvement = ((post_score - pre_score) / pre_score) * 100
                    else:
                        improvement = 0
                    improvement_values.append(improvement)

                improvement_avg_percentage = int(round(sum(improvement_values) / len(improvement_values))) if improvement_values else None

                # Trainer isimlerini topla
                trainers = group.trainers.all()
                trainer_names = [trainer.get_full_name() or trainer.username for trainer in trainers]
                trainers_display = ", ".join(trainer_names) if trainer_names else "—"

                rows.append({
                    "customer_name": group.customer.name if group.customer else "",
                    "program_name": ps.program.name,
                    "training_group_name": group.name,
                    "trainers": trainers_display,
                    "scenario_name": scenario.name,
                    "release_date": ps.release_date.strftime('%d.%m.%Y') if ps.release_date else "—",
                    "training_date": ps.training_date.strftime('%d.%m.%Y') if ps.training_date else "—",
                    "close_date": ps.close_date.strftime('%d.%m.%Y') if ps.close_date else "—",
                    "pre_completed": pre_completed,
                    "pre_avg": pre_avg,
                    "pre_attempt_avg": pre_attempt_avg,
                    "post_completed": post_completed,
                    "post_avg": post_avg,
                    "post_attempt_avg": post_attempt_avg,
                    "improvement_avg_percentage": improvement_avg_percentage,
                    "total_users": total_users,
                })

        # Build XLSX with merged headers (admin görünümündeki gibi)
        wb = Workbook()
        ws = wb.active
        ws.title = "Müşteri Raporu"

        # Header rows
        header_top = [
            "Müşteri", "Program", "Eğitim Grubu", "Eğitmenler", "Senaryo",
            "Yayın Tarihi", "Eğitim Tarihi", "Kapanış Tarihi",
            "Toplam Kullanıcı",
            "ÖN DEĞERLENDİRME", "", "",
            "SON DEĞERLENDİRME", "", "",
            "Gelişim (%)",
        ]
        header_bottom = [
            "", "", "", "", "",
            "", "", "",
            "",
            "Tamamlanma", "Ort.Puan", "Ort.Deneme",
            "Tamamlanma", "Ort.Puan", "Ort.Deneme",
            "",
        ]

        ws.append(header_top)
        ws.append(header_bottom)

        # Merge single columns down
        merge_single_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 16]
        for col in merge_single_cols:
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

        # Merge grouped headers
        ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)   # Ön değerlendirme
        ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=15)  # Son değerlendirme

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row_idx in (1, 2):
            for col_idx in range(1, 17):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

        # Data rows
        for row in rows:
            ws.append([
                row["customer_name"],
                row["program_name"],
                row["training_group_name"],
                row["trainers"],
                row["scenario_name"],
                row["release_date"],
                row["training_date"],
                row["close_date"],
                row["total_users"],
                f'{row["pre_completed"]}/{row["total_users"]}',
                row["pre_avg"],
                f'{row["pre_attempt_avg"]:.2f}',
                f'{row["post_completed"]}/{row["total_users"]}',
                row["post_avg"],
                f'{row["post_attempt_avg"]:.2f}',
                row["improvement_avg_percentage"] if row["improvement_avg_percentage"] is not None else "",
            ])
            # Trend arrows with color (green up triangle, red down triangle)
            r_idx = ws.max_row
            improvement_cell = ws.cell(row=r_idx, column=16)
            val = row["improvement_avg_percentage"]
            if val is None:
                improvement_cell.value = "-"
            else:
                if val > 0:
                    arrow, color = "▲", "008000"  # green
                    improvement_cell.value = f"{arrow} {val}%"
                elif val < 0:
                    arrow, color = "▼", "C00000"  # red
                    improvement_cell.value = f"{arrow} {val}%"
                else:
                    improvement_cell.value = "-"
                    color = "000000"  # black for neutral
                improvement_cell.font = Font(bold=True, color=color)
                improvement_cell.alignment = center

        # Apply borders/alignment to data rows
        for r in range(3, ws.max_row + 1):
            for c in range(1, 17):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                if c >= 10 and c <= 15:
                    cell.alignment = center

        # Auto-width (basic)
        # Auto-width (basic) — skip merged placeholder cells
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                cell = row[0]
                # Ignore empty merged placeholders
                if cell.value is None:
                    continue
                value_len = len(str(cell.value))
                if value_len > max_length:
                    max_length = value_len
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="musteri_raporu.xlsx"'
        return response

    def autocomplete_training_groups(self, request, object_id):
        """Select2-compatible autocomplete endpoint for training groups"""
        from django.http import JsonResponse
        
        customer = self.get_object(request, object_id)
        if customer is None:
            return JsonResponse({'results': [], 'pagination': {'more': False}})

        search_term = request.GET.get('term', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = 20

        queryset = TrainingGroup.objects.filter(customer=customer).select_related('program')
        
        if search_term:
            queryset = queryset.filter(name__icontains=search_term)
        
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        training_groups = queryset.order_by('name')[start:end]
        
        results = [{'id': tg.id, 'text': tg.name} for tg in training_groups]
        
        return JsonResponse({
            'results': results,
            'pagination': {'more': end < total}
        })

    def autocomplete_programs(self, request, object_id):
        """Select2-compatible autocomplete endpoint for programs"""
        from django.http import JsonResponse
        
        customer = self.get_object(request, object_id)
        if customer is None:
            return JsonResponse({'results': [], 'pagination': {'more': False}})

        search_term = request.GET.get('term', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = 20

        queryset = Program.objects.filter(traininggroup__customer=customer).distinct()
        
        if search_term:
            queryset = queryset.filter(name__icontains=search_term)
        
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        programs = queryset.order_by('name')[start:end]
        
        results = [{'id': p.id, 'text': p.name} for p in programs]
        
        return JsonResponse({
            'results': results,
            'pagination': {'more': end < total}
        })

    def autocomplete_program_scenarios(self, request, object_id):
        """Select2-compatible autocomplete endpoint for program scenarios"""
        from django.http import JsonResponse
        
        customer = self.get_object(request, object_id)
        if customer is None:
            return JsonResponse({'results': [], 'pagination': {'more': False}})

        search_term = request.GET.get('term', '').strip()
        program_id = request.GET.get('program_id', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = 20

        queryset = ProgramScenario.objects.filter(
            program__traininggroup__customer=customer
        ).select_related('scenario', 'program').distinct()
        
        if program_id:
            queryset = queryset.filter(program_id=program_id)
        
        if search_term:
            queryset = queryset.filter(
                Q(scenario__name__icontains=search_term) | 
                Q(program__name__icontains=search_term)
            )
        
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        program_scenarios = queryset.order_by('program__name', 'scenario__name')[start:end]
        
        results = [
            {
                'id': ps.id,
                'text': f"{ps.program.name} | {ps.scenario.name}"
            }
            for ps in program_scenarios
        ]
        
        return JsonResponse({
            'results': results,
            'pagination': {'more': end < total}
        })


class GroupUserInline(admin.TabularInline):
    model = GroupUser
    extra = 0
    can_delete = False
    readonly_fields = ("user", "progress_display", "total_score_display", "scenarios_display")
    fields = ("user", "progress_display", "total_score_display", "scenarios_display")
    verbose_name_plural = "Eğitim Grubu Katılımcıları"

    def get_extra(self, request, obj=None, **kwargs):
        extra = super().get_extra(request, obj, **kwargs)
        if obj:
            # Store the training group object for use in scenarios_display
            self._training_group_obj = obj
        return extra

    def progress_display(self, obj):
        return format_html(
            '''
            <div style="
                background-color: rgba(200, 200, 200, 0.3);
                width: 80px;
                height: 16px;
                border-radius: 3px;
                overflow: hidden;
                border: 1px solid rgba(120, 120, 120, 0.5);
            ">
                <div style="
                    width: {}%;
                    background: linear-gradient(to right, #4caf50, #388e3c);
                    height: 100%;
                    font-size: 11px;
                    color: white;
                    text-align: center;
                ">
                    {}%
                </div>
            </div>
            ''',
            obj.progress,
            obj.progress
        )
    progress_display.short_description = "İlerleme"

    def scenarios_display(self, obj):
        program_scenarios = obj.training_group.program.programscenario_set.all().order_by('release_date')
        user = obj.user
        group = obj.training_group

        if not program_scenarios.exists():
            return "-"

        rows = ""
        for ps in program_scenarios:
            scenario_name = ps.scenario.name
            release_date = ps.release_date.strftime('%d.%m.%Y') if ps.release_date else "Yayın tarihi yok"

            # En güncel raporları al
            pre_report = (
                EvaluationReport.objects
                .filter(user=user, training_group=group, scenario=ps.scenario, type="pre")
                .order_by('-created_at')
                .first()
            )
            post_report = (
                EvaluationReport.objects
                .filter(user=user, training_group=group, scenario=ps.scenario, type="post")
                .order_by('-created_at')
                .first()
            )

            # Puan ve tarih
            pre_score = pre_report.score if pre_report else ""
            post_score = post_report.score if post_report else ""
            pre_date = timezone.localtime(pre_report.created_at).strftime('%d.%m.%Y') if pre_report else ""
            post_date = timezone.localtime(post_report.created_at).strftime('%d.%m.%Y') if post_report else ""

            # Ortalama puan
            pre_avg_score = float(pre_report.average_score) if pre_report and pre_report.average_score is not None else ""
            post_avg_score = float(post_report.average_score) if post_report and post_report.average_score is not None else ""

            # Deneme sayısı
            pre_attempt = pre_report.attempt_count if pre_report else ""
            post_attempt = post_report.attempt_count if post_report else ""

            # Linkler
            pre_score_html = str(pre_score)
            pre_date_html = pre_date
            post_score_html = str(post_score)
            post_date_html = post_date
            if pre_report:
                pre_link = reverse("admin:my_app_evaluationreport_view_report", args=[pre_report.id])
                pre_score_html = f'<a href="{pre_link}" target="_blank" title="Raporu görüntüle">{pre_score}</a>'
                pre_date_html = f'<a href="{pre_link}" target="_blank" title="Raporu görüntüle">{pre_date}</a>'
            if post_report:
                post_link = reverse("admin:my_app_evaluationreport_view_report", args=[post_report.id])
                post_score_html = f'<a href="{post_link}" target="_blank" title="Raporu görüntüle">{post_score}</a>'
                post_date_html = f'<a href="{post_link}" target="_blank" title="Raporu görüntüle">{post_date}</a>'

            # Gelişim yüzdesi (post - pre)
            improvement_cell = ""
            if pre_report and post_report:
                pre_score_val = pre_score or 0
                post_score_val = post_score or 0
                if pre_score_val > 0:
                    improvement_percentage = int(round(((post_score_val - pre_score_val) / pre_score_val) * 100, 0))
                else:
                    improvement_percentage = 0
                trend_color = "#2e7d32" if improvement_percentage >= 0 else "#c62828"
                trend_arrow = "▲" if improvement_percentage >= 0 else "▼"
                improvement_cell = f"<span style=\"color:{trend_color};font-weight:600;\">{trend_arrow} {improvement_percentage}%</span>"

            # Ortalama puan formatı
            pre_avg_score_display = f"{pre_avg_score:.2f}" if pre_avg_score != "" else ""
            post_avg_score_display = f"{post_avg_score:.2f}" if post_avg_score != "" else ""

            rows += f"""
                <tr>
                    <td title=\"{scenario_name}\" style=\"border: 1px solid #ccc; padding: 4px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-size: 12px;\">{scenario_name}</td>
                    <td title=\"{release_date}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: left; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-variant-numeric: tabular-nums; font-size: 11px; vertical-align: middle;\">{release_date}</td>
                    <td title=\"{pre_score}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 12px;\">{pre_score_html}</td>
                    <td title=\"{pre_avg_score_display}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 11px; color: #666;\">{pre_avg_score_display}</td>
                    <td title=\"{pre_date}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; font-variant-numeric: tabular-nums; font-size: 11px; vertical-align: middle;\">{pre_date_html}</td>
                    <td title=\"{pre_attempt}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 11px;\">{pre_attempt}</td>
                    <td title=\"{post_score}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 12px;\">{post_score_html}</td>
                    <td title=\"{post_avg_score_display}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 11px; color: #666;\">{post_avg_score_display}</td>
                    <td title=\"{post_date}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; font-variant-numeric: tabular-nums; font-size: 11px; vertical-align: middle;\">{post_date_html}</td>
                    <td title=\"{post_attempt}\" style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 11px;\">{post_attempt}</td>
                    <td style=\"border: 1px solid #ccc; padding: 2px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; vertical-align: middle; font-size: 11px;\">{improvement_cell}</td>
                </tr>
            """

        help_icon_url = static('admin/img/icon-unknown.svg')

                # Add export button if training group object is available
        export_button = ""
        if hasattr(self, '_training_group_obj') and self._training_group_obj:
            export_url = reverse(
                f"admin:{self._training_group_obj._meta.app_label}_{self._training_group_obj._meta.model_name}_export_all_participants_xlsx",
                args=[self._training_group_obj.pk]
            )
            export_button = f'<div style="float: right; margin: 0; display: inline-block; margin-top: -6px;"><ul class="object-tools" style="margin: 0; font-size: 9px;"><li><a href="{export_url}" style="padding: 3px 10px;" title="Tüm Katılımcıların Detay Raporunu Dışa Aktar">Dışa Aktar</a></li></ul></div>'

        html = f"""
            <div style=\"width: 700px; vertical-align: middle; text-align: left;\">
                <details style=\"display: block;\">
                    <summary style=\"cursor: pointer; font-weight: bold; margin: 0;\">📈 Senaryo Detaylarını Göster {export_button}</summary>
                    <div style=\"margin-top: 8px;\">
                        <table style=\"border-collapse: collapse; width: 100%; font-size: 12px; table-layout: fixed;\">
                            <thead>
                                <tr style=\"background-color: #f5f5f5;\">
                                    <th rowspan=\"2\" style=\"border: 1px solid #ccc; padding: 4px; vertical-align: middle; width: 21%; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-size: 10px;\">Senaryo</th>
                                    <th rowspan=\"2\" style=\"border: 1px solid #ccc; padding: 4px; vertical-align: middle; width: 9%; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-variant-numeric: tabular-nums; font-size: 10px;\">Yayın Tarihi</th>
                                    <th colspan=\"4\" style=\"border: 1px solid #ccc; padding: 4px; text-align: center; width: 32%; white-space: nowrap; vertical-align: middle; font-size: 10px;\">Ön Değerlendirme</th>
                                    <th colspan=\"4\" style=\"border: 1px solid #ccc; padding: 4px; text-align: center; width: 32%; white-space: nowrap; vertical-align: middle; font-size: 10px;\">Son Değerlendirme</th>
                                    <th rowspan=\"2\" style=\"border: 1px solid #ccc; padding: 4px; text-align: center; vertical-align: middle; width: 11%; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-size: 10px;\">Gelişim (%)</th>
                                </tr>
                                <tr style=\"background-color: #f9f9f9;\">
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 10px;\">Puan</th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 9px;\">Ort. <img class=\"help-tooltip\" src=\"{help_icon_url}\" alt=\"?\" title=\"Kullanıcının bu senaryo için tüm denemelerinin ortalaması\" style=\"width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;\"/></th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 30%; white-space: nowrap; font-variant-numeric: tabular-nums; vertical-align: middle; text-align: center; font-size: 10px;\">Tarih</th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 9px;\">Deneme <img class=\"help-tooltip\" src=\"{help_icon_url}\" alt=\"?\" title=\"Katılımcının bu senaryodaki deneme sayısı\" style=\"width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;\"/></th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 10px;\">Puan</th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 9px;\">Ort. <img class=\"help-tooltip\" src=\"{help_icon_url}\" alt=\"?\" title=\"Kullanıcının bu senaryo için tüm denemelerinin ortalaması\" style=\"width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;\"/></th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 30%; white-space: nowrap; font-variant-numeric: tabular-nums; vertical-align: middle; text-align: center; font-size: 10px;\">Tarih</th>
                                    <th style=\"border: 1px solid #ccc; padding: 4px 1px; width: 20px; min-width: 20px; max-width: 20px; white-space: nowrap; vertical-align: middle; text-align: center; font-size: 9px;\">Deneme <img class=\"help-tooltip\" src=\"{help_icon_url}\" alt=\"?\" title=\"Katılımcının bu senaryodaki deneme sayısı\" style=\"width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;\"/></th>
                                </tr>
                            </thead>
                            <tbody>
                                {rows}
                            </tbody>
                        </table>
                    </div>
                </details>
            </div>
        """
        return format_html(html)
    scenarios_display.short_description = "Senaryolar (Tarih & Rapor Durumu)"

    def total_score_display(self, obj):
        return obj.total_score if obj.progress == 100 else "-"
    # Add help tooltip to header, sourcing text from the model field help_text
    help_icon_url = static('admin/img/icon-unknown.svg')
    help_text = GroupUser._meta.get_field('total_score').help_text
    total_score_display.short_description = format_html(
        'Ortalama Puan <img class="help-tooltip" src="{}" alt="?" title="{}" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/>',
        help_icon_url,
        help_text,
    )

class TrainingGroupCustomerFilter(AutocompleteFilter):
    title = 'Müşteri'
    field_name = 'customer'
    
class TrainingGroupProgramFilter(AutocompleteFilter):
    title = 'Program'
    field_name = 'program'

class TrainingGroupTrainersFilter(AutocompleteFilter):
    title = 'Eğitmenler'
    field_name = 'trainers'
    
def build_training_group_summary_rows(training_group):
    """Return (total_users, rows) for the summary table.

    rows include: scenario_name, pre_* and post_* metrics, improvement_avg_percentage.
    Averages match the admin UI; improvement uses only users with both scores.
    """
    program_scenarios = training_group.program.programscenario_set.select_related('scenario').all().order_by('release_date')
    group_users = training_group.groupuser_set.select_related('user')
    total_users = group_users.count()
    if not program_scenarios.exists() or total_users == 0:
        return total_users, []

    rows = []
    # Sadece katılımcıların (users) raporlarını dahil et, trainer'ları hariç tut
    participant_users = training_group.users.all()
    
    for ps in program_scenarios:
        scenario = ps.scenario

        pre_qs = EvaluationReport.objects.filter(
            training_group=training_group,
            scenario=scenario,
            type="pre",
            user__in=participant_users,
        )
        pre_completed = pre_qs.count()
        pre_avg = int(round(pre_qs.aggregate(avg=Avg("score"))["avg"] or 0, 2))
        pre_attempt_sum = pre_qs.aggregate(total=Sum("attempt_count"))["total"] or 0
        pre_attempt_avg = round(pre_attempt_sum / total_users, 2) if total_users > 0 else 0

        post_qs = EvaluationReport.objects.filter(
            training_group=training_group,
            scenario=scenario,
            type="post",
            user__in=participant_users,
        )
        post_completed = post_qs.count()
        post_avg = int(round(post_qs.aggregate(avg=Avg("score"))["avg"] or 0, 2))
        post_attempt_sum = post_qs.aggregate(total=Sum("attempt_count"))["total"] or 0
        post_attempt_avg = round(post_attempt_sum / total_users, 2) if total_users > 0 else 0

        pre_scores_by_user = dict(pre_qs.values_list("user_id", "score"))
        post_scores_by_user = dict(post_qs.values_list("user_id", "score"))
        improvement_values = []
        for user_id, pre_score in pre_scores_by_user.items():
            post_score = post_scores_by_user.get(user_id)
            if post_score is None:
                continue
            if pre_score and pre_score > 0:
                improvement = ((post_score - pre_score) / pre_score) * 100
            else:
                improvement = 0
            improvement_values.append(improvement)

        if improvement_values:
            improvement_avg_percentage = int(round(sum(improvement_values) / len(improvement_values)))
        else:
            improvement_avg_percentage = None

        rows.append({
            "scenario_name": scenario.name,
            "pre_completed": pre_completed,
            "pre_avg": pre_avg,
            "pre_attempt_avg": pre_attempt_avg,
            "post_completed": post_completed,
            "post_avg": post_avg,
            "post_attempt_avg": post_attempt_avg,
            "improvement_avg_percentage": improvement_avg_percentage,
        })

    return total_users, rows
    
@admin.register(TrainingGroup)
class TrainingGroupAdmin(admin.ModelAdmin):
    list_display = ("name_display", "program", "trainers_display", "customer", "progress_bar", "group_date")
    readonly_fields = ("progress_bar", "report_summary")
    list_filter = (TrainingGroupCustomerFilter, TrainingGroupProgramFilter, TrainingGroupTrainersFilter, "group_date")
    autocomplete_fields = ['customer', 'program']
    search_fields = ("name", "program__name", "customer__name")
    ordering = ['-id']

    def name_display(self, obj):
        # Inline styling to give the name column more breathing room in list view
        return format_html('<span style="display:inline-block; min-width: 260px;">{}</span>', obj.name)
    name_display.admin_order_field = "name"
    name_display.short_description = "Grup İsmi"

    def trainers_display(self, obj):
        trainers = obj.trainers.all()
        if not trainers.exists():
            return "-"
        trainer_names = [trainer.get_full_name() or trainer.username for trainer in trainers]
        return ", ".join(trainer_names)
    trainers_display.short_description = "Eğitmenler"

    def progress_bar(self, obj):
        return format_html(
            '''
            <div style="
                background-color: rgba(200, 200, 200, 0.3);
                width: 100px;
                height: 16px;
                border-radius: 3px;
                overflow: hidden;
                border: 1px solid rgba(120, 120, 120, 0.5);
            ">
                <div style="
                    width: {}%;
                    background: linear-gradient(to right, #4caf50, #388e3c);
                    height: 100%;
                    font-size: 11px;
                    color: white;
                    text-align: center;
                ">
                    {}%
                </div>
            </div>
            ''',
            obj.progress,
            obj.progress
        )
    progress_bar.short_description = "İlerleme %"

    def get_form(self, request, obj=None, change=False, **kwargs):
        if obj is not None:
            kwargs['form'] = TrainingGroupAdminForm
        return super().get_form(request, obj, change=change, **kwargs)

    def get_fields(self, request, obj=None):
        fields = ["name", "program", "customer", "group_date"]
        if obj:
            fields.append("users")  # sadece düzenleme ekranında göster
            fields.append("trainers")  # eğitmenler
            fields.append("report_summary")
        return fields
    
    def get_inlines(self, request, obj=None):
        if obj:
            return [GroupUserInline]
        return []
     
    def response_add(self, request, obj, post_url_continue=None):
        self.message_user(
            request,
            "🎯 Eğitim grubu başarıyla oluşturuldu. Şimdi kullanıcıları ekleyebilirsiniz.",
            level=messages.INFO
        )
        change_url = reverse("admin:my_app_traininggroup_change", args=[obj.pk])
        return HttpResponseRedirect(change_url)

    def get_urls(self):
        """Add change-view XLSX export endpoints."""
        urls = super().get_urls()
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        custom = [
            path(
                '<int:object_id>/export-summary-xlsx/',
                self.admin_site.admin_view(self.export_summary_xlsx),
                name=f'{app_label}_{model_name}_export_summary_xlsx',
            ),
            path(
                '<int:object_id>/export-all-participants-xlsx/',
                self.admin_site.admin_view(self.export_all_participants_xlsx),
                name=f'{app_label}_{model_name}_export_all_participants_xlsx',
            ),
        ]
        return custom + urls

    #Grup Senaryo Raporu Özeti
    def report_summary(self, obj):
        """Render summary table + right-aligned export button (desktop), top-left on mobile."""
        total_users, row_dicts = build_training_group_summary_rows(obj)
        if not row_dicts:
            return "Senaryo veya kullanıcı verisi bulunamadı."

        rows = ""
        for r in row_dicts:
            if r['improvement_avg_percentage'] is not None:
                trend_color = "#2e7d32" if r['improvement_avg_percentage'] >= 0 else "#c62828"
                trend_arrow = "▲" if r['improvement_avg_percentage'] >= 0 else "▼"
                improvement_cell = f"<span style=\"color:{trend_color};font-weight:600;\">{trend_arrow} {r['improvement_avg_percentage']}%</span>"
            else:
                improvement_cell = "-"

            rows += f"""
                <tr>
                    <td style=\"border: 1px solid #ccc; padding: 4px; white-space: nowrap;\">{r['scenario_name']}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['pre_completed']}/{total_users}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['pre_avg']}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['pre_attempt_avg']}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['post_completed']}/{total_users}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['post_avg']}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{r['post_attempt_avg']}</td>
                    <td style=\"border: 1px solid #ccc; text-align: center; padding: 4px; white-space: nowrap;\">{improvement_cell}</td>
                </tr>
            """

        help_icon_url = static('admin/img/icon-unknown.svg')
        export_url = reverse(
            f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_export_summary_xlsx",
            args=[obj.pk],
        )

        # Namespaced minimal CSS; avoids bleeding into admin styles.
        html = f"""
        <div style="min-width: 300px;">
            <div style="overflow-x: auto; overflow-y: hidden; -webkit-overflow-scrolling: touch; width: 100%; scrollbar-gutter: stable both-edges;">
            <details>
                <summary style="cursor: pointer; font-weight: bold; margin-bottom: 8px;">📋 Raporu Göster</summary>
                <style>
                .tg-summary-row {{ display: flex; align-items: flex-start; gap: 12px; }}
                .tg-summary-row .tg-object-tools {{ margin: 0; min-width: max-content; }}
                .tg-object-tools .object-tools {{ margin: 0; float: none; }}
                .tg-object-tools .object-tools li {{ margin: 0; }}
                .tg-object-tools .object-tools li a {{
                    white-space: nowrap !important;
                    word-break: keep-all;
                    overflow-wrap: normal;
                    display: inline-flex;
                    align-items: center;
                }}
                @media (max-width: 768px) {{
                    .tg-summary-row {{ flex-direction: column; }}
                    .tg-summary-row .tg-object-tools {{ order: -1; align-self: flex-start; margin: 4px 0; }}
                    .tg-object-tools .object-tools {{ float: none; }}
                }}
                </style>
                <div class=\"tg-summary-row\"> 
                    <div style=\"overflow-x: auto; overflow-y: hidden; -webkit-overflow-scrolling: touch; touch-action: pan-x; width: 100%; display: block; padding-bottom: 8px; scrollbar-gutter: stable both-edges;\">
                    <div style="display: inline-block; white-space: nowrap;">
                    <table style="border-collapse: collapse; font-size: 13px; width: auto; display: table; table-layout: auto; white-space: nowrap;">
                        <thead>
                            <tr style="background-color: #f5f5f5;">
                                <th rowspan="2" style="border: 1px solid #ccc; padding: 4px; vertical-align: middle; white-space: nowrap;">Senaryo</th>
                                <th colspan="3" style="border: 1px solid #ccc; text-align: center; white-space: nowrap;">Ön Değerlendirme</th>
                                <th colspan="3" style="border: 1px solid #ccc; text-align: center; white-space: nowrap;">Son Değerlendirme</th>
                                <th rowspan="2" style="border: 1px solid #ccc; padding: 4px; vertical-align: middle; white-space: nowrap;">Gelişim (%)</th>
                            </tr>
                            <tr style="background-color: #f9f9f9;">
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Tamamlanma <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Tamamlayan Katılımcı/Toplam Katılımcı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Ort.Puan <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Katılımcıların Ort.Puanı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Ort.Deneme <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Katılımcı Başına Deneme Sayısı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Tamamlanma <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Tamamlayan Katılımcı/Toplam Katılımcı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Ort.Puan <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Katılımcıların Ort.Puanı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                                <th style="border: 1px solid #ccc; padding: 4px; white-space: nowrap;">Ort.Deneme <img class="help-tooltip" src="{help_icon_url}" alt="?" title="Katılımcı Başına Deneme Sayısı" style="width:10px;height:10px;vertical-align:middle;position:relative;top:-2px;"/></th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows}
                        </tbody>
                    </table>
                        </div>
                    </div>
                    <div class=\"tg-object-tools\">
                        <ul class=\"object-tools\" style=\"margin:0;\">
                            <li><a href=\"{export_url}\">Dışa Aktar</a></li>
                        </ul>
                    </div>
                </div>
            </details>
            </div>
        </div>
        """
        return mark_safe(html)

    report_summary.short_description = "Eğitim Grubu Özet Raporu"

    def export_summary_xlsx(self, request, object_id):
        """Generate XLSX matching the report_summary table; includes meta block and merged headers."""
        obj = self.get_object(request, object_id)
        if obj is None:
            return HttpResponse(status=404)

        total_users, row_dicts = build_training_group_summary_rows(obj)
        if not row_dicts:
            return HttpResponse("Dışa aktarılacak veri bulunamadı.", status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Özet Rapor"

        # Dynamic columns; no hard-coded counts
        pre_headers = ["Tamamlanma", "Ort.Puan", "Ort.Deneme"]
        post_headers = ["Tamamlanma", "Ort.Puan", "Ort.Deneme"]
        second_row_headers = pre_headers + post_headers
        total_columns = 1 + len(second_row_headers) + 1  # Scenario + detailed + Improvement

        # Meta information block (top of sheet)
        participant_count = obj.groupuser_set.count()
        progress_int = int(round(float(obj.progress or 0)))
        progress_fraction = progress_int / 100.0  # numeric, formatted as percentage later
        avg_total_score_val = GroupUser.objects.filter(training_group=obj).aggregate(avg=Avg('total_score'))['avg'] or 0.0
        avg_total_score_num = float(avg_total_score_val)
        meta_rows = [
            ("Grup İsmi", obj.name or ""),
            ("Program", getattr(obj.program, 'name', '') or ""),
            ("Müşteri", getattr(obj.customer, 'name', '') or ""),
            ("Eğitim Tarihi", obj.group_date.strftime('%d.%m.%Y') if getattr(obj, 'group_date', None) else ""),
            ("Katılımcı Sayısı", participant_count),
            ("İlerleme %", progress_fraction),
            ("Grup Ort.Puanı", avg_total_score_num),
        ]
        # Rapor Tarihi sabit konumda: G1 (7. sütun), tarih H1 (8. sütun)
        ws.cell(row=1, column=7, value="Rapor Tarihi").font = Font(bold=True)
        ws.cell(row=1, column=8, value=timezone.localtime().strftime('%d.%m.%Y'))
        for idx, (label, value) in enumerate(meta_rows, start=1):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            # Merge B:F alanı
            ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
            val_cell = ws.cell(row=idx, column=2, value=value)
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            if label == "İlerleme %":
                val_cell.number_format = '0%'
            elif label == "Grup Ort.Puanı":
                val_cell.number_format = '0.00'

        # Define visible borders
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_border = Border(left=medium, right=medium, top=medium, bottom=medium)
        label_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
        for r in range(1, len(meta_rows) + 1):
            for c in range(1, total_columns + 1):
                ws.cell(row=r, column=c).border = cell_border
            # Label column styling and uniform row height
            lab = ws.cell(row=r, column=1)
            lab.fill = label_fill
            ws.row_dimensions[r].height = 18
        # Rapor Tarihi başlık hücresi stili (G1-H1)
        ws.cell(row=1, column=7).fill = label_fill
        ws.cell(row=1, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=8).alignment = Alignment(horizontal="right", vertical="center")

        # Table headers start after an empty spacer row below the meta block
        table_start_row = len(meta_rows) + 2
        ws.cell(row=table_start_row, column=1, value="Senaryo")
        ws.merge_cells(start_row=table_start_row, start_column=1, end_row=table_start_row + 1, end_column=1)

        # Ön Değerlendirme başlık hücresi 2..pre_end
        pre_start = 2
        pre_end = pre_start + len(pre_headers) - 1
        ws.cell(row=table_start_row, column=pre_start, value="Ön Değerlendirme")
        ws.merge_cells(start_row=table_start_row, start_column=pre_start, end_row=table_start_row, end_column=pre_end)

        # Son Değerlendirme başlık hücresi pre_end+1 .. post_end
        post_start = pre_end + 1
        post_end = post_start + len(post_headers) - 1
        ws.cell(row=table_start_row, column=post_start, value="Son Değerlendirme")
        ws.merge_cells(start_row=table_start_row, start_column=post_start, end_row=table_start_row, end_column=post_end)

        # Improvement header (last column == total_columns)
        improvement_col = post_end + 1
        ws.cell(row=table_start_row, column=improvement_col, value="Gelişim (%)")
        ws.merge_cells(start_row=table_start_row, start_column=improvement_col, end_row=table_start_row + 1, end_column=improvement_col)

        # Second row headers under merged blocks
        for i, title in enumerate(second_row_headers, start=2):
            ws.cell(row=table_start_row + 1, column=i, value=title)

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", start_color="F5F5F5", end_color="F5F5F5")
        for r in (table_start_row, table_start_row + 1):
            for c in range(1, total_columns + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.alignment = center
                cell.border = header_border
                cell.fill = header_fill

        # Body rows
        excel_row = table_start_row + 2
        for r in row_dicts:
            ws.cell(excel_row, 1, r['scenario_name']).alignment = Alignment(vertical="center")
            ws.cell(excel_row, 2, f"{r['pre_completed']}/{total_users}").alignment = center
            ws.cell(excel_row, 3, r['pre_avg']).alignment = center
            ws.cell(excel_row, 4, r['pre_attempt_avg']).alignment = center
            pre_written_end = 1 + len(pre_headers)
            ws.cell(excel_row, pre_written_end + 1, f"{r['post_completed']}/{total_users}").alignment = center
            ws.cell(excel_row, pre_written_end + 2, r['post_avg']).alignment = center
            ws.cell(excel_row, pre_written_end + 3, r['post_attempt_avg']).alignment = center
            improvement_text = f"{r['improvement_avg_percentage']}%" if r['improvement_avg_percentage'] is not None else "-"
            cell_impr = ws.cell(excel_row, improvement_col, improvement_text)
            cell_impr.alignment = center
            # Optional color cue similar to UI arrow color
            if r['improvement_avg_percentage'] is not None:
                if r['improvement_avg_percentage'] >= 0:
                    cell_impr.font = Font(color="008000")
                else:
                    cell_impr.font = Font(color="CC0000")
            for c in range(1, total_columns + 1):
                ws.cell(excel_row, c).border = cell_border
            excel_row += 1

        # Column widths (safely sized according to column count)
        widths = [22] + [14, 12, 12] + [14, 12, 12] + [12]
        for i in range(1, total_columns + 1):
            w = widths[i - 1] if i - 1 < len(widths) else 12
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header after second header row
        ws.freeze_panes = f"A{table_start_row + 2}"

        # Return as response
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"egitim_grubu_ozet_raporu_{obj.pk}.xlsx"
        response = HttpResponse(
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_all_participants_xlsx(self, request, object_id):
        """Generate XLSX for all participants data matching the scenarios_display table format."""
        obj = self.get_object(request, object_id)
        if obj is None:
            return HttpResponse(status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "Tüm Katılımcılar"

        # Meta information block
        participant_count = obj.groupuser_set.count()
        progress_int = int(round(float(obj.progress or 0)))
        progress_fraction = progress_int / 100.0
        avg_total_score_val = GroupUser.objects.filter(training_group=obj).aggregate(avg=Avg('total_score'))['avg'] or 0.0
        avg_total_score_num = float(avg_total_score_val)
        
        meta_rows = [
            ("Grup İsmi", obj.name or ""),
            ("Program", getattr(obj.program, 'name', '') or ""),
            ("Müşteri", getattr(obj.customer, 'name', '') or ""),
            ("Eğitim Tarihi", obj.group_date.strftime('%d.%m.%Y') if getattr(obj, 'group_date', None) else ""),
            ("Katılımcı Sayısı", participant_count),
            ("İlerleme %", progress_fraction),
            ("Grup Ort.Puanı", avg_total_score_num),
        ]
        
        # Rapor Tarihi sağ üstte
        ws.cell(row=1, column=8, value="Rapor Tarihi").font = Font(bold=True)
        ws.cell(row=1, column=9, value=timezone.localtime().strftime('%d.%m.%Y'))
        
        for idx, (label, value) in enumerate(meta_rows, start=1):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            # Merge B:G alanı (7 sütun)
            ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=7)
            val_cell = ws.cell(row=idx, column=2, value=value)
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            if label == "İlerleme %":
                val_cell.number_format = '0%'
            elif label == "Grup Ort.Puanı":
                val_cell.number_format = '0.00'

        # Border styling for meta block
        thin = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        label_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
        
        for r in range(1, len(meta_rows) + 1):
            for c in range(1, 10):  # 9 sütun (A:I)
                ws.cell(row=r, column=c).border = cell_border
            # Label column styling
            lab = ws.cell(row=r, column=1)
            lab.fill = label_fill
            ws.row_dimensions[r].height = 18
        # Rapor Tarihi styling
        ws.cell(row=1, column=8).fill = label_fill
        ws.cell(row=1, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=9).alignment = Alignment(horizontal="right", vertical="center")

        # Table headers start after meta block
        table_start_row = len(meta_rows) + 2
        
        # Main headers with merged cells
        ws.cell(row=table_start_row, column=1, value="Katılımcı")
        ws.merge_cells(start_row=table_start_row, start_column=1, end_row=table_start_row + 1, end_column=1)
        
        ws.cell(row=table_start_row, column=2, value="İlerleme")
        ws.merge_cells(start_row=table_start_row, start_column=2, end_row=table_start_row + 1, end_column=2)
        
        ws.cell(row=table_start_row, column=3, value="Ort. Puan")
        ws.merge_cells(start_row=table_start_row, start_column=3, end_row=table_start_row + 1, end_column=3)
        
        ws.cell(row=table_start_row, column=4, value="Senaryo")
        ws.merge_cells(start_row=table_start_row, start_column=4, end_row=table_start_row + 1, end_column=4)
        
        ws.cell(row=table_start_row, column=5, value="Yayın Tarihi")
        ws.merge_cells(start_row=table_start_row, start_column=5, end_row=table_start_row + 1, end_column=5)
        
        ws.cell(row=table_start_row, column=6, value="Ön Değerlendirme")
        ws.merge_cells(start_row=table_start_row, start_column=6, end_row=table_start_row, end_column=8)
        
        ws.cell(row=table_start_row, column=9, value="Son Değerlendirme")
        ws.merge_cells(start_row=table_start_row, start_column=9, end_row=table_start_row, end_column=11)
        
        ws.cell(row=table_start_row, column=12, value="Gelişim (%)")
        ws.merge_cells(start_row=table_start_row, start_column=12, end_row=table_start_row + 1, end_column=12)

        # Second row headers
        second_headers = ["Puan", "Tarih", "Deneme", "Puan", "Tarih", "Deneme"]
        for i, title in enumerate(second_headers, start=6):
            ws.cell(row=table_start_row + 1, column=i, value=title)

        # Header styling
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", start_color="F5F5F5", end_color="F5F5F5")
        
        for r in (table_start_row, table_start_row + 1):
            for c in range(1, 13):  # 12 sütun
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.alignment = center
                cell.border = header_border
                cell.fill = header_fill

        # Body rows - get all participants and scenarios
        program_scenarios = obj.program.programscenario_set.select_related('scenario').all().order_by('release_date')
        group_users = obj.groupuser_set.select_related('user').all()
        excel_row = table_start_row + 2
        
        for group_user in group_users:
            user = group_user.user
            user_progress = float(group_user.progress or 0) / 100.0
            user_total_score = float(group_user.total_score or 0)
            scenario_count = program_scenarios.count()
            
            # First row for this user
            first_user_row = excel_row
            
            for ps in program_scenarios:
                scenario_name = ps.scenario.name
                release_date = ps.release_date.strftime('%d.%m.%Y') if ps.release_date else "Yayın tarihi yok"

                # Get reports for this user and scenario
                pre_report = (
                    EvaluationReport.objects
                    .filter(user=user, training_group=obj, scenario=ps.scenario, type="pre")
                    .order_by('-created_at')
                    .first()
                )
                post_report = (
                    EvaluationReport.objects
                    .filter(user=user, training_group=obj, scenario=ps.scenario, type="post")
                    .order_by('-created_at')
                    .first()
                )

                # Data values
                pre_score = pre_report.score if pre_report else ""
                post_score = post_report.score if post_report else ""
                pre_date = timezone.localtime(pre_report.created_at).strftime('%d.%m.%Y') if pre_report else ""
                post_date = timezone.localtime(post_report.created_at).strftime('%d.%m.%Y') if post_report else ""
                pre_attempt = pre_report.attempt_count if pre_report else ""
                post_attempt = post_report.attempt_count if post_report else ""

                # Improvement calculation
                improvement_text = ""
                if pre_report and post_report and pre_score and pre_score > 0:
                    improvement_percentage = int(round(((post_score - pre_score) / pre_score) * 100, 0))
                    trend_arrow = "▲" if improvement_percentage >= 0 else "▼"
                    improvement_text = f"{trend_arrow} {improvement_percentage}%"

                # Write row data
                ws.cell(excel_row, 1, user.get_full_name() or user.username)
                ws.cell(excel_row, 2, user_progress)
                ws.cell(excel_row, 3, user_total_score)
                ws.cell(excel_row, 4, scenario_name)
                ws.cell(excel_row, 5, release_date)
                ws.cell(excel_row, 6, pre_score)
                ws.cell(excel_row, 7, pre_date)
                ws.cell(excel_row, 8, pre_attempt)
                ws.cell(excel_row, 9, post_score)
                ws.cell(excel_row, 10, post_date)
                ws.cell(excel_row, 11, post_attempt)
                ws.cell(excel_row, 12, improvement_text)

                # Apply borders and alignment
                for c in range(1, 13):
                    cell = ws.cell(excel_row, c)
                    cell.border = cell_border
                    if c in [6, 8, 9, 11]:  # Score and attempt columns
                        cell.alignment = center
                    elif c == 12:  # Improvement column
                        cell.alignment = center
                        if improvement_text and "▲" in improvement_text:
                            cell.font = Font(color="008000")
                        elif improvement_text and "▼" in improvement_text:
                            cell.font = Font(color="CC0000")

                excel_row += 1

            # Merge user info cells across all scenarios for this user
            if scenario_count > 1:
                # Katılımcı ismi merge
                ws.merge_cells(start_row=first_user_row, start_column=1, end_row=first_user_row + scenario_count - 1, end_column=1)
                
                # İlerleme merge
                ws.merge_cells(start_row=first_user_row, start_column=2, end_row=first_user_row + scenario_count - 1, end_column=2)
                
                # Ort. Puan merge
                ws.merge_cells(start_row=first_user_row, start_column=3, end_row=first_user_row + scenario_count - 1, end_column=3)
                
                # Set number formats and alignment for merged cells
                participant_cell = ws.cell(first_user_row, 1)
                participant_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                progress_cell = ws.cell(first_user_row, 2)
                progress_cell.number_format = '0%'
                progress_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                score_cell = ws.cell(first_user_row, 3)
                score_cell.number_format = '0.00'
                score_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths - optimized for better readability
        widths = [25, 12, 12, 25, 15, 12, 12, 12, 12, 12, 12, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header
        ws.freeze_panes = f"A{table_start_row + 2}"

        # Return response
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"egitim_grubu_katilimci_detay_raporu_{obj.pk}.xlsx"
        response = HttpResponse(
            stream.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    
class ProgramScenarioInline(admin.TabularInline):
    model = ProgramScenario
    formset = ProgramScenarioInlineFormSet
    extra = 1
    autocomplete_fields = ['scenario']
    verbose_name = "Senaryo"
    fields = ['scenario', 'release_date', 'training_date', 'close_date', 'weight_percentage', 'max_attempts']
    verbose_name_plural = "Eğitim Senaryoları"

class ProgramScenariosFilter(AutocompleteFilter):
    title = 'Eğitim Senaryoları'
    field_name = 'scenarios'

@admin.register(Program)
class ProgramAdmin(admin.ModelAdmin):
    list_display = ("name", "created_at")
    search_fields = ("name",)
    inlines = [ProgramScenarioInline]
    list_filter = ['created_at', ProgramScenariosFilter]

class EvaluationReportCustomerFilter(AutocompleteFilter):
    title = 'Müşteri'
    field_name = 'customer'

class EvaluationReportScenarioFilter(AutocompleteFilter):
    title = 'Senaryo'
    field_name = 'scenario'

class EvaluationReportTrainingGroupFilter(AutocompleteFilter):
    title = 'Eğitim Grubu'
    field_name = 'training_group'

class EvaluationReportUserFilter(AutocompleteFilter):
    title = 'Kullanıcı'
    field_name = 'user'

@admin.register(EvaluationReport)
class EvaluationReportAdmin(admin.ModelAdmin):
    list_display = ('user', 'training_group', 'scenario', 'score', 'created_at')
    search_fields = ('user__username', 'scenario__name', 'training_group__name')
    list_filter = (
        EvaluationReportTrainingGroupFilter,
        EvaluationReportScenarioFilter,
        EvaluationReportCustomerFilter,
        EvaluationReportUserFilter,
        'created_at'
    )
    #readonly_fields = ['type', 'user', 'customer', 'training_group', 'scenario', 'report', 'score', 'created_at', 'attempt_count']

    def get_urls(self):
        urls = super().get_urls()
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        custom = [
            path(
                '<int:object_id>/view-report/',
                self.admin_site.admin_view(self.view_report),
                name=f'{app_label}_{model_name}_view_report',
            ),
            path(
                '<int:object_id>/export-pdf/',
                self.admin_site.admin_view(self.export_pdf),
                name=f'{app_label}_{model_name}_export_pdf',
            ),
        ]
        return custom + urls

    def view_report(self, request, object_id):
        """Report'u markdown formatında görüntüle"""
        from django.shortcuts import get_object_or_404
        from django.template.response import TemplateResponse
        import markdown

        report_obj = get_object_or_404(EvaluationReport, pk=object_id)
        
        # Markdown'ı HTML'e çevir
        html_content = markdown.markdown(
            report_obj.report,
            extensions=['fenced_code', 'tables', 'nl2br']
        )
        
        context = {
            **self.admin_site.each_context(request),
            'report': report_obj,
            'html_content': html_content,
            'title': f'Değerlendirme Raporu - {report_obj.user.get_full_name() or report_obj.user.username}',
            'opts': self.model._meta,
        }
        
        return TemplateResponse(request, 'admin/my_app/evaluationreport/view_report.html', context)

    def export_pdf(self, request, object_id):
        """Report'u PDF olarak export et"""
        from django.shortcuts import get_object_or_404
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        import markdown
        from weasyprint import HTML, CSS
        from weasyprint.text.fonts import FontConfiguration

        report_obj = get_object_or_404(EvaluationReport, pk=object_id)
        
        # Markdown'ı HTML'e çevir
        html_content = markdown.markdown(
            report_obj.report,
            extensions=['fenced_code', 'tables', 'nl2br']
        )
        
        # Template context
        context = {
            'report': report_obj,
            'html_content': html_content,
        }
        
        # HTML string oluştur
        html_string = render_to_string('admin/my_app/evaluationreport/pdf_template.html', context)
        
        # PDF oluştur
        font_config = FontConfiguration()
        css = CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 12pt;
                line-height: 1.6;
            }
            h1, h2, h3 {
                color: #333;
                margin-top: 1em;
                margin-bottom: 0.5em;
            }
            table {
                border-collapse: collapse;
                width: 100%;
                margin: 1em 0;
            }
            table th, table td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            table th {
                background-color: #f5f5f5;
            }
            pre {
                background-color: #f5f5f5;
                padding: 10px;
                border-radius: 4px;
                overflow-x: auto;
            }
            code {
                background-color: #f5f5f5;
                padding: 2px 4px;
                border-radius: 3px;
            }
        ''', font_config=font_config)
        
        html_doc = HTML(string=html_string)
        pdf_file = html_doc.write_pdf(stylesheets=[css], font_config=font_config)
        
        # Response oluştur
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'degerlendirme_raporu_{report_obj.id}_{report_obj.user.username}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

# Custom filtre tanımı
class CustomerListFilter(admin.SimpleListFilter):
    title = _('Müşteri')  # sağ sütunda görünen başlık
    parameter_name = 'customer'  # URL parametresi
    template = 'admin/customer_filter.html'

    def lookups(self, request, model_admin):
        # Filtre listesi (müşteri isimleri)
        return [(c.id, c.name) for c in Customer.objects.all()]

    def queryset(self, request, queryset):
        # Filtre uygulandığında yapılacak sorgu
        if self.value():
            return queryset.filter(customer__id=self.value())  # reverse M2M
        return queryset

#import_export
class UserResource(resources.ModelResource):
    customer = fields.Field(column_name='musteri', readonly=True)

    def dehydrate_customer(self, user):
        return user.customer.first().name if user.customer.exists() else "-"
    
    def before_save_instance(self, instance, row, **kwargs):
        # E-posta geçerli olup olmadığını kontrol et
        try:
            validate_email(instance.email)
        except ValidationError:
            raise ValidationError(f"⚠️ '{instance.email}' geçerli bir e-posta değil.")
        
        if not instance.username:
            instance.username = instance.email
        
        # Email ve username eşleşiyor mu?
        if instance.username != instance.email:
            raise ValidationError("⚠️ Username ve email alanları aynı olmalıdır.")

        if not instance.pk:
            # Random şifre oluştur
            password = User.objects.make_random_password()
            instance.set_password(password)

            # Şifreyi geçici olarak loglamak istersen (opsiyonel):
            print(f"Yeni kullanıcı: {instance.email} | Şifre: {password}")
            # Veya güvenli log dosyasına yazabilirsin

    class Meta:
        model = User
        import_id_fields = ['username']
        fields = ('username', 'email', 'first_name', 'last_name', 'is_active', 'customer')

def send_welcome_email_manual(user):
    """Manuel olarak kullanıcıya hoş geldiniz e-postası gönderir (signals.py'deki mantığı kullanır)."""
    email = user.username  # çünkü username=email olmalı

    if not email or not user.is_active:
        return False

    if not user.email:
        user.email = user.username
        user.save(update_fields=["email"])

    form = PasswordResetForm({'email': email})
    if form.is_valid():
        form.save(
            request=None,
            use_https=not settings.DEBUG,
            domain_override="127.0.0.1:8000" if settings.DEBUG else "ml-simulator.onrender.com",
            email_template_name='registration/welcome_email.html',
            subject_template_name='registration/welcome_subject.txt',
        )
        return True
    return False

# User admin'i genişletiyoruz
class CustomUserAdmin(ImportExportModelAdmin, DefaultUserAdmin):
    add_form = CustomUserCreationForm
    form = CustomUserChangeForm
    resource_classes = (UserResource,)
    list_filter = DefaultUserAdmin.list_filter + (CustomerListFilter,)
    list_display = ('username', 'first_name', 'last_name', 'is_staff', 'get_customers')
    actions = ['send_welcome_email_action']
    change_list_template = 'admin/auth/user/change_list.html'

    def get_customers(self, obj):
        return ", ".join([c.name for c in obj.customer.all() if c.name]) or "-"
    get_customers.short_description = "Müşteri"

    def get_search_results(self, request, queryset, search_term):
        """
        Limit user autocomplete results for TrainingGroup.trainers to
        'Dale Carnegie Akademi' users only. Other autocomplete usages stay unchanged.
        """
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        if (
            request.GET.get('app_label') == 'my_app'
            and request.GET.get('model_name') == 'traininggroup'
            and request.GET.get('field_name') == 'trainers'
        ):
            try:
                dale_carnegie = Customer.objects.get(name="Dale Carnegie Akademi")
                queryset = queryset.filter(customer=dale_carnegie)
            except Customer.DoesNotExist:
                queryset = queryset.none()
        return queryset, use_distinct

    def changelist_view(self, request, extra_context=None):
        """Online kullanıcı bilgilerini changelist sayfasına ekle."""
        extra_context = extra_context or {}
        
        # Online kullanıcıları al
        online_users = get_online_users()
        online_count = len(online_users)
        
        # Online kullanıcı listesi HTML'i oluştur
        online_users_html = ""
        if online_users:
            users_list = []
            for user_data in online_users[:10]:  # İlk 10 kullanıcıyı göster
                full_name = f"{user_data.get('first_name', '')} {user_data.get('last_name', '')}".strip()
                display_name = full_name if full_name else user_data.get('username', 'N/A')
                users_list.append(f"<li>{display_name} ({user_data.get('email', 'N/A')})</li>")
            
            if len(online_users) > 10:
                users_list.append(f"<li><em>... ve {len(online_users) - 10} kullanıcı daha</em></li>")
            
            online_users_html = f"<ul style='margin: 0; padding-left: 20px;'>{''.join(users_list)}</ul>"
        else:
            online_users_html = "<p style='margin: 0; color: #666;'>Şu anda online kullanıcı yok.</p>"
        
        extra_context['online_users_count'] = online_count
        extra_context['online_users_list'] = mark_safe(online_users_html)
        
        return super().changelist_view(request, extra_context)

    def send_welcome_email_action(self, request, queryset):
        """Seçili kullanıcılara hoş geldiniz e-postası gönderir (toplu gönderim)."""
        success_count = 0
        error_count = 0
        
        for user in queryset:
            if send_welcome_email_manual(user):
                success_count += 1
            else:
                error_count += 1
        
        if success_count > 0:
            self.message_user(
                request,
                f'✅ {success_count} kullanıcıya hoş geldiniz e-postası gönderildi.',
                level=messages.SUCCESS
            )
        if error_count > 0:
            self.message_user(
                request,
                f'❌ {error_count} kullanıcıya e-posta gönderilemedi (aktif olmayan veya geçersiz e-posta).',
                level=messages.WARNING
            )
    
    send_welcome_email_action.short_description = "Seçili kullanıcılara hoş geldiniz e-postası gönder"

    def get_urls(self):
        urls = super().get_urls()
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        custom = [
            path(
                '<int:object_id>/send-welcome-email/',
                self.admin_site.admin_view(self.send_welcome_email_view),
                name=f'{app_label}_{model_name}_send_welcome_email',
            ),
        ]
        return custom + urls

    def send_welcome_email_view(self, request, object_id):
        """Tek kullanıcı için hoş geldiniz e-postası gönderir."""
        user = self.get_object(request, object_id)
        if user is None:
            self.message_user(request, 'Kullanıcı bulunamadı.', level=messages.ERROR)
            return HttpResponseRedirect(reverse('admin:auth_user_changelist'))
        
        if send_welcome_email_manual(user):
            self.message_user(
                request,
                f'✅ Hoş geldiniz e-postası {user.username} adresine gönderildi.',
                level=messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                f'❌ E-posta gönderilemedi. Kullanıcı aktif değil veya e-posta geçersiz.',
                level=messages.ERROR
            )
        
        return HttpResponseRedirect(reverse('admin:auth_user_change', args=[object_id]))

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_send_email_button'] = True
        return super().change_view(request, object_id, form_url, extra_context)

# Default UserAdmin'i kaldırıp kendi versiyonumuzu ekliyoruz
admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)
